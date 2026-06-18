"""
Converte os resultados do scraper OLX (CSV) em uma planilha Excel formatada.
Rode este script DEPOIS de rodar o olx.py.

Uso:
    .\\olx_scraper\\venv\\Scripts\\python.exe gerar_excel.py
"""
import csv
import re
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def limpar_preco(preco_str):
    """Converte 'R$ 500.000' para o numero 500000."""
    if not preco_str:
        return None
    numeros = re.sub(r'[^\d]', '', preco_str)
    if numeros:
        return int(numeros)
    return None


def converter_timestamp(ts_str):
    """Converte timestamp Unix para data legivel."""
    if not ts_str:
        return ''
    try:
        ts = int(ts_str)
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M')
    except (ValueError, OSError):
        return ts_str


def gerar_excel(arquivo_csv):
    if not os.path.exists(arquivo_csv):
        print(f"Arquivo {arquivo_csv} não encontrado.")
        return None

    arquivo_excel = arquivo_csv.replace('.csv', '.xlsx')
    print(f"Gerando Excel para: {arquivo_csv}")

    # --- Le o CSV ---
    with open(arquivo_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        linhas = list(reader)
        campos = reader.fieldnames

    if not linhas:
        print("Nenhum dado encontrado no CSV.")
        return

    print(f"Lidos {len(linhas)} anuncios do CSV.")

    # --- Cria a planilha ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Anuncios OLX"

    # Cores e estilos
    COR_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    FONTE_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    FONTE_NORMAL = Font(name="Calibri", size=10)
    FONTE_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")
    BORDA = Border(
        bottom=Side(style="thin", color="D9D9D9")
    )
    ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center")
    ALINHAMENTO_ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ALINHAMENTO_DIREITA = Alignment(horizontal="right", vertical="center")

    COR_ZEBRA_1 = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    COR_ZEBRA_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Mapeamento de colunas para nomes bonitos em PT-BR
    NOMES_COLUNAS = {
        'title': 'Titulo',
        'price': 'Preco (R$)',
        'location': 'Localizacao',
        'city': 'Cidade',
        'neighborhood': 'Bairro',
        'state': 'Estado',
        'url': 'Link do Anuncio',
        'list_id': 'ID do Anuncio',
        'date': 'Data de Publicacao',
        'description': 'Descricao',
        'author': 'Anunciante',
        'phone': 'Telefone',
        'detail_status': 'Status Detalhe',
    }

    # --- Cabecalho ---
    for col_idx, campo in enumerate(campos, 1):
        cell = ws.cell(row=1, column=col_idx, value=NOMES_COLUNAS.get(campo, campo))
        cell.font = FONTE_HEADER
        cell.fill = COR_HEADER
        cell.alignment = ALINHAMENTO_CENTRO
        cell.border = BORDA

    # --- Dados ---
    for row_idx, linha in enumerate(linhas, 2):
        cor_linha = COR_ZEBRA_1 if row_idx % 2 == 0 else COR_ZEBRA_2

        for col_idx, campo in enumerate(campos, 1):
            valor = linha.get(campo, '')

            # Formatacoes especiais por coluna
            if campo == 'price':
                preco_num = limpar_preco(valor)
                cell = ws.cell(row=row_idx, column=col_idx, value=preco_num)
                cell.number_format = '#,##0'
                cell.alignment = ALINHAMENTO_DIREITA
            elif campo == 'date':
                cell = ws.cell(row=row_idx, column=col_idx, value=converter_timestamp(valor))
                cell.alignment = ALINHAMENTO_CENTRO
            elif campo == 'url':
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                if valor:
                    cell.hyperlink = valor
                    cell.font = FONTE_LINK
                cell.alignment = ALINHAMENTO_ESQUERDA
            elif campo == 'description':
                # Limita descricao a 500 caracteres para nao explodir a planilha
                texto = valor[:500] + '...' if len(valor) > 500 else valor
                cell = ws.cell(row=row_idx, column=col_idx, value=texto)
                cell.alignment = ALINHAMENTO_ESQUERDA
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = ALINHAMENTO_ESQUERDA

            cell.fill = cor_linha
            cell.border = BORDA
            if campo != 'url':
                cell.font = FONTE_NORMAL

    # --- Largura automatica das colunas ---
    LARGURAS = {
        'title': 55,
        'price': 15,
        'location': 35,
        'city': 18,
        'neighborhood': 20,
        'state': 18,
        'url': 40,
        'list_id': 15,
        'date': 20,
        'description': 60,
        'author': 25,
        'phone': 18,
        'detail_status': 15,
    }

    for col_idx, campo in enumerate(campos, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = LARGURAS.get(campo, 20)

    # --- Congela o cabecalho ---
    ws.freeze_panes = 'A2'

    # --- Adiciona filtros automaticos ---
    ws.auto_filter.ref = ws.dimensions

    # --- Altura da linha do cabecalho ---
    ws.row_dimensions[1].height = 30

    # --- Salva ---
    wb.save(arquivo_excel)
    print(f"\nPlanilha Excel gerada com sucesso!")
    print(f"Arquivo: {arquivo_excel}")
    print(f"Total de linhas: {len(linhas)}")
    
    return arquivo_excel


if __name__ == '__main__':
    # Teste rápido se rodar direto
    arquivos_csv = glob.glob('resultados_*.csv')
    if arquivos_csv:
        arquivo_recente = max(arquivos_csv, key=os.path.getmtime)
        gerar_excel(arquivo_recente)
    else:
        print("Nenhum arquivo CSV encontrado.")
